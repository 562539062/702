from tkinter import messagebox
import functions as fc
import openpyxl
from openpyxl.utils import get_column_letter
import numpy as np
import settings as s
from openpyxl.styles import Alignment
import os



def gen_excel(savepath,runfiles,select_items,select_point_index):
    ###生成对应需求的excel初始表格
    wb=openpyxl.Workbook()
    ws=wb.active
    sumname=["Full_",'Peak_',"Valley_","average_", "meaningful_","max_", "syn", "high","wave"]
    excelname=['全幅值','峰值','谷值','平均值','有义值','最大值','合成成分','高频成分','波浪成分']
    name1=[]
    name2=[]
    name3=[]
    value1=[]
    value2=[]
    value3=[]
    savename=runfiles[0][0:4]+'.'+runfiles[0][4:6]+'.'+runfiles[0][6:8]+'-'+runfiles[-1][0:4]+'.'+runfiles[-1][4:6]+'.'+runfiles[-1][6:8]+'__'
    index=0
    for flag in select_items[0:3]:
        if flag:
            value1.append(sumname[index])
            name1.append(excelname[index])
            savename+=excelname[index]
        index += 1
    savename+='-'
    for flag in select_items[3:6]:
        if flag:
            value2.append(sumname[index])
            name2.append(excelname[index])
            savename+=excelname[index]
        index += 1
    savename+='-'
    for flag in select_items[6:9]:
        if flag:
            value3.append(sumname[index])
            name3.append(excelname[index])
            savename+=excelname[index]
        index += 1
    select_pointname=[]  #表示选择的特殊值
    for each1 in value1:
        for each2 in value2:
            for each3 in value3:
                select_pointname.append(each1 + each2 + each3)
    ##excel初始表格
    pointnum=0 #表示表格内第几个测点，以此来判断所有初始化的值应该填在哪个位置
    l1_s=len(name1)*len(name2)*len(name3)   #line1_space 每个测点占多少列
    l2_s=len(name2)*len(name3)              #line2_space 每个测点中的一类特殊值占多少列
    l3_s=len(name3)                         #line3_space 每个测点中的二类特殊值占多少列
    for i in select_point_index:
        ws.cell(row=1,column=pointnum*l1_s+2).value=s.Signame[i]
        ws.merge_cells(get_column_letter(pointnum*l1_s+2)+'1:'+get_column_letter((pointnum+1)*l1_s+1)+'1')
        for j in range(len(name1)):
            ws.cell(row=2,column=pointnum*l1_s+2+j*l2_s).value=name1[j]
            ws.merge_cells(get_column_letter(pointnum*l1_s+2+j*l2_s)+'2:'+get_column_letter(pointnum*l1_s+2+(j+1)*l2_s-1)+'2')
            for k in range(len(name2)):
                ws.cell(row=3,column=pointnum*l1_s+2+j*l2_s+k*l3_s).value=name2[k]
                ws.merge_cells(get_column_letter(pointnum*l1_s+2+j*l2_s+k*l3_s)+'3:'+get_column_letter(pointnum*l1_s+2+j*l2_s+(k+1)*l3_s-1)+'3')
                for l in range(len(name3)):
                    ws.cell(row=4,column=pointnum*l1_s+2+j*l2_s+k*l3_s+l).value=name3[l]
        pointnum+=1
        
    alignment_center = Alignment(horizontal='center', vertical='center')
    for i in ws:
        for j in i:
            j.alignment = alignment_center
            
    ###将数据填入       
    row_num=5
    if_file_complete=1
    sumS=fc.getfileseg(runfiles)
    for fn in range(len(runfiles)):
        ftime=runfiles[fn][0:4]+'-'+runfiles[fn][4:6]+'-'+runfiles[fn][6:8]+' '
        for i in range(sumS[fn]):  #设置时间
            time=ftime+str((int(runfiles[fn][8:10])+int((30*i+int(runfiles[fn][10:12]))/60))%24)+':'+str((int(runfiles[fn][10:12])+(i%2)*30)%60)
            ws.cell(row=row_num+i,column=1).value=time
        pointnum=0
        print(runfiles[fn],row_num)
        for j in select_point_index:
            for i in range(len(select_pointname)):
                if not os.path.exists(savepath+'plot_data/'+s.Signame[j]+'/'+runfiles[fn]+'/'+select_pointname[i]+'.npy'):
                    if_file_complete=0
                else:
                    pointdata=np.load(savepath+'plot_data/'+s.Signame[j]+'/'+runfiles[fn]+'/'+select_pointname[i]+'.npy')
                    for k in range(len(pointdata)):
                        ws.cell(row=row_num+k,column=pointnum*l1_s+2+i).value=pointdata[k]
            pointnum+=1
        row_num+=sumS[fn]
    if if_file_complete==0:
        messagebox.showinfo(title='warning', message='Files are missing, and the obtained excel is incomplete\n'\
                        'Please complete the running operation for the selected time period first')
    messagebox.showinfo(title='show information', message='excel show')
    if len(select_point_index)==1:
        path=savepath+'excel/'+s.Signame[select_point_index[0]]+'/'+savename+'.xlsx'
        fc.fun_save(path,'xlsx',wb)
        os.startfile(path)
    else:
        fc.fun_save(savepath+'excel/汇总.xlsx','xlsx',wb)
        os.startfile(savepath+'excel/汇总.xlsx')
        
        

