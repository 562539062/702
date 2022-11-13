from pprint import pp
import shutil
from statistics import mean
import scipy
import win32api
import win32con
import math
import os
import numpy as np
from scipy.fftpack import fft,ifft
from ctypes import sizeof
from numpy.linalg import pinv
import scipy.io as scio
from distutils.log import error
from re import X
from turtle import color
from settings import *
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import functions as fc
from json import load
from matplotlib.pyplot import MultipleLocator
from timeit import timeit
from datetime import datetime
from openpyxl.styles import Alignment


def loadSignal20201129(filename, pathstr, segment_length = 720000, padding =True): 
#从文件中提取数据，采用单元数组存储大部分数据,一般数据存储姿态传感器数据
# Sig 除姿态之外的信号
# Sig1 姿态传感器输出信号
# SigN 除姿态之外的通道总数
# SigZ 除姿态之外的通道总数
# Nright 有效数据序号
# Nerror 记录缺少记录的测点序号
# 姿态传感器50Hz采样
# 最后修改时间2020/11/29
# 修改内容,姿态传感器调整为50Hz
# 采用频率一样，不再分别读取姿态和应力数据
    signal_len = len(filename) #signal type number: standard 164
    
    Nerror = []
    Nright = []
    Sig = []
    Len = np.zeros(signal_len)
    #liu:164是标准信号数，159可能是因为当天数据有丢失，具体原因目前不明
    #search exist file id
    for loc in range(len(filename)):
        name = os.path.join(pathstr,filename[loc][0][0])
        if os.path.exists(name) and os.path.getsize(pathstr) and os.path.getsize(name):#liu:返回是否存在该列数据;wang:检查是否为空文件夹以及是否缺数据
            fid = open(name,'rb')
            SampleFrequency = np.fromfile(fid, np.int32)[0] #read frequency
            
            fid = open(name,'rb')
            temp = np.fromfile(fid, np.float32)[1:] #read signal 精度有问题
            #数据可能差1秒
            
            #liu:这里不考虑时间对齐数据，从位置1开始重新存入，时间不对应，可能是问题
            #generate length and signal list
            if segment_length-len(temp)>0 and padding:
                Sig.append(np.concatenate([temp.squeeze(),np.zeros(segment_length-len(temp))],0))
            else:
                Sig.append(temp.squeeze())
            Len[loc] = len(temp)
            Nright.append(loc) #add right id
        else:
            Sig.append(np.zeros(segment_length))
        
            Nerror.append(loc) # add false id
            #liu:貌似是在用数组循环计数
        #fid.close()
        
    if not len(set(Nright+Nerror)) == signal_len:
        raise("file length error")
    ##修复MO1-Yaw信号
    # temp=Sig1(:,1);
    # temp(temp<-100)=temp(temp<-100)+360;
    # plot(Sig1(:,1)), hold on, plot(temp)
    # Sigl(:,1)=temp;
    #return Sig, np.array(Len), np.array(Nright)
    return Sig, np.array(Len), np.array(Nright), Nerror

#transfer date time to the target format
def transfer_date_name(ttemp):
    return ttemp[:4] + '/' + ttemp[4:6] + '/' + ttemp[6:8] + ' ' + ttemp[8:10] + ':' + ttemp[10:12] + ':' + ttemp[12:]
    
def Acc2Disp(Acc,SFreq,FF):
    #单列加速度信号转换成单列位移信号
    #开始编辑日期2020/1/2
    #最后编辑日期2020/1/2
    #编辑者 徐春
    #FF=0.0005;#低频截止频率
    dt=1/SFreq
    Acc=Sig_detrend(Acc, SFreq)
        #消除零飘
        #Acc=Acc-mean(Acc)
    Velo=np.cumsum(Acc)*dt;#获取速度信号
        #plot(Velo),title("Velo"), pause
    Velo=Sig_detrend(Velo, SFreq);#消除零飘
        #plot(Velo),title("消除零飘 Velo") , pause
    Velo=fftfilter(Velo, SFreq,FF, SFreq/7); #提取分析信号
        #plot(Velo),title("消除零飘 提取分析信号Velo") , pause
    Disp=np.cumsum(Velo)*dt;#获取位移信号
        #plot(Disp),title("Disp获取位移信号"), pause
    Disp=Sig_detrend(Disp,SFreq);#消除零飘
    Disp=fftfilter(Disp, SFreq, FF, SFreq/7);#提取分析信号
        #plot (Disp),title("Disp 提取分析信号"), pause
    return Disp


def Sig_detrend(x, Freq):
    #消除信号零飘，傅里叶滤波前处理
    #单列信号
                                                
    xlen=np.size(x)
    x=np.array(x)
    x=x.reshape(xlen,1)
    N=np.shape(x)[0]
    bp=[i for i in range(300*Freq-1,xlen-300*Freq,300*Freq)]#以6分钟数据进行线性趋势消除(该处是否需要-1，数组从0 开始)
    bp=[0]+bp+[N-1]
    bp=np.array(bp)
    lbp=np.size(bp)
    a=np.zeros((N,lbp))
    a[:N,0]=np.array(range(1,N+1))/N
    for k in range(1,lbp-1):
        M=N-bp[k]-1
        a[bp[k]+1:,k]=np.array(range(1,M+1))/M;
    a[:N,lbp-1]=1
    y=x-np.dot(a,(np.dot(pinv(a),x)))
    y=np.array(y)
    y=y.flatten()
    return y

# def Sig_detrend(x, Freq):
    #消除信号零飘，傅里叶滤波前处理
    #单列信号                                            
    xlen=np.size(x)
    tp=[i for i in range(300*Freq-1,xlen-1,300*Freq)]#以6分钟数据进行线性趋势消除(该处是否需要-1，数组从0 开始)
    y=scipy.signal.detrend(x,type="linear",bp=tp)
    # print(tp)
    # y=np.empty(xlen)
    # for i in tp:
    #     y[i:i+2]=scipy.signal.detrend(x[i:i+2],type="linear")
    return y

def fftfilter(uy,uFreq,u1,u2):
    #Signal filtering
#uy Signal column
#uFreq采样频率
#ul u2 Filtering Freq.
#Output Filtered Signal
#EDIT TIME 2008/10/16
#Lat modified Time 2013/11/07
    uy=np.array(uy)
    try:
        [Line,column]=np.shape(uy)
        
    except:
        Line=len(uy)
        column=1
        uy=uy.reshape(Line,1)
    if Line<column:
            uy=uy.T
            [Line,column]=np.shape(uy)
#将数据调整,多少信号就多少列

    uLowFreq=min(u1,u2)
    uHighFreq=max(u1,u2)
    if uHighFreq*2>uFreq:
        win32api.MessageBox(0,'滤波高频频率应小于最高采样频率',win32con.MB_ICONWARNING)
        uHighFreq=uFreq/2.02
    uLen=Line#数据长度
    uIndexLow=math.floor(uLen*uLowFreq/uFreq);
    ulndexL=uLen+2-uIndexLow;

    uIndexHigh=math.ceil(uLen*uHighFreq/uFreq);
    uIndexH=uLen+2-uIndexHigh;
    y=np.zeros(np.shape(uy));
    
    for PP in range(column):
        uFFT=fft(uy[:,PP])
        temindex=list(range(uIndexLow))+list(range(uIndexHigh-1,uIndexH))+list(range(ulndexL-1,np.shape(uFFT)[0]))
        uFFT[temindex]=0
        y[:,PP]=np.real(ifft(uFFT))
    if column==1:
        y=y.flatten()
    return y
    

def De_Sample(Data,uFreq,Low,High):
# 抽取分析信号中低通,高通，合成子样、
# Data数据（以列形式排列)
# Freq, Low,High 采样频率,低频滤波频率,高频滤波频率
# edit by xuchun 2010/08/11

    x = fftfilter(Data,uFreq, Low, High)
    nm_temp = [len(Data),len(Data[0])] # 获取取矩阵Data的维度
    NM=min(nm_temp)
    Data=np.array(Data)
    Result_Sample=np.empty(NM,dtype="object")
    for k in range(NM): # 信号个数
        # 过零点统计
        s = Data[:,k] # 分析信号（预先滤过波)
        sx = x[:,k] # 低频波浪弯矩信号
        sh = s - sx # 高频船体结构响应信号
        h =  np.diff(np.sign(sx))
        inn = [i for i,a in enumerate(h) if a == 2]
        inn = list(inn)
        # 信号抽取
        kk = len(inn)
        Result_Sample[k] = np.shape((kk, 10))
        if kk >= 2:
            Result=np.zeros((kk-1,10))
            Result[:,9]=list(map(lambda a:a/uFreq, np.diff(inn[:kk])))
            for j in range(kk-1):
                # k= 41时报错，inn取不到
                Result[j][0]=np.max(np.array(sx[(inn[j]):(inn[j+1]+1)])) # 低频峰值
                Result[j][1]=np.min(np.array(sx[(inn[j]):(inn[j+1]+1)])) # 低频谷值
                Result[j][2]=Result[j][0]-Result[j][1] # 低频峰谷值
                Result[j][3]=np.max(np.array(sh[(inn[j]):(inn[j+1]+1)])) # 高频峰值
                Result[j][4]=np.min(np.array(sh[(inn[j]):(inn[j+1]+1)])) # 高频谷值
                Result[j][5]=Result[j][3]-Result[j][4] # 高频峰谷值
                Result[j][6]=np.max(np.array(s[(inn[j]):(inn[j+1]+1)])) # 合成峰值
                Result[j][7]=np.min(np.array(s[(inn[j]):(inn[j+1]+1)])) # 合成谷值
                Result[j][8]=Result[j][6]-Result[j][7] # 合成峰谷值
            Result_Sample[k] = Result
    return Result_Sample

def Sta_irregular(Result_Sample):
# 统计分析不规则波信号中峰值、谷值、峰谷值及其特征值
# 输出平均值、三一值、最大值
###########################
# 第一行:低频统计值;第二行:高频统计值;第三行:合成弯矩统计值
###########################
#                           不规则波
#     中拱（峰值）            中垂（谷值）            全幅值               周期
# 平均值 三一值 最大值   平均值 三一值 最大值   平均值 三一值 最大值   平均遭遇周期
# edit by xuchun 2008/12/12
# Last modified on 4, June, 2008

    NM= 1
    result=np.zeros(28)
    for k in range(NM): # 信号个数
        Result = np.abs(Result_Sample)
        Result = np.sort(Result, axis=0)[::-1]
        Result=np.array(Result)
        kk = len(Result)
        if kk<=1:
            continue
        for num in range(3):
            result[num*3] = np.mean(Result[:,num])
            result[num*3+1] = np.mean(Result[0:round(kk/3-1),num])
            result[num*3+2] = np.max(Result[:,num])
            result[9+num*3] = np.mean(Result[:,num+3])
            result[9+num*3+1] = np.mean(Result[0:round(kk/3-1),num+3])
            result[9+num*3+2] = np.max(Result[:,num+3])
            result[18+num*3] = np.mean(Result[:,num+6])
            result[18+num*3+1] = np.mean(Result[0:round(kk/3-1),num+6])
            result[18+num*3+2] = np.max(Result[:,num+6])
        result[27] = np.mean(Result[:,9])
    return result



def fun_save(path,dtype,data):
    checkpath=os.path.abspath(os.path.join(os.path.dirname(path)))
    isExistpath=os.path.exists(checkpath)
    if not isExistpath:
        os.makedirs(checkpath)
    if dtype=='mat':
        scio.savemat(path,mdict=data)
    if dtype=='img':
        data.savefig(path,dpi=600)
    if dtype=='xlsx':
        data.save(path)
    if dtype=='npy':
        np.save(path,data)
        
# def preinit(savepath):  #最新运行时所做的初始化
#     if os.path.exists(savepath):
#         shutil.rmtree(savepath)
    
    
def createtimeline(runfiles):
    timeline=[]
    for ffilename in runfiles:
        file_hour=int(ffilename[8:-4])
        file_min=int(ffilename[10:-2])
        for Segment in range(1,sumSegment+1):
            timeline.append(ffilename[0:4]+'-'+ffilename[4:6]+'-'+ffilename[6:8]+' '+str((file_hour+int((30*(Segment-1)+file_min)/60))%24)+':'+str((file_min+((Segment-1)%2)*30)%60))
    return timeline



            
def pddeal(savepath,ffilename,tempplotdata):   #由于多进程乱序运行，保存所得plot数据不按顺序保存，将分别保存的数据按顺序保存入一个文件中
    for j in index_3:
        for i in range(len(pointname)):
            path=savepath+'plot_data/'+str(Signame[j][0])[2:-2]+'/'+ffilename+'/'+pointname[i]+'.npy'
            sumdata=[]
            for k in range(sumSegment):
                sumdata.append(tempplotdata[j*len(pointname)*8+i*sumSegment+k])
            fun_save(path,'npy',sumdata)
    
                
def signal1(path,ffilename,savepath):
    tstr = transfer_date_name(ffilename)
    pathstr = os.path.join(path, ffilename)  # 文件路径
    [Sig, Len, Nright, Nerror]= loadSignal20201129(filename, pathstr, segment_length = 720000, padding=False)
    # 将文件夹名称转为Python可以识别的日期字符串，转换成数值写入结构数组中
    # liu: 保存名称更换格式，后续其它文件可能对名称有依赖
    ###########################################

    #  单向应力传感器数据修正
    #  2019/9/26以后，后期数据无需修正
    #  for j=1:10
    #  	Sig{j,1}=Sig{j,1}*Coef(j); # 转换单向应力结果
    #  end

    #  信号滤波 
    #  Sig=Sigtemp;
    #  Sigtemp=Sig;
    #  for j=[1:145 149:151 155:157 159]
    #  x=ELchange(Sig{j}); #消除单点跳变
    #  Sig{j}=Sig_detrend(x,50); #消除零飘 
    #  	Sig{j}=Sig_detrend(Sig{j},50);
    #  	plot([Sigtemp{j} Sig{j}])
    #  	legend(Signame(j))
    #  	pause
    #  end


    # M2无信号，Z2和Z5无信号
    # 信号显示1

    if True:
        dates = datetime.strptime(tstr,'%Y/%m/%d %H:%M:%S') # 将字符串转化为 matplotlib的datetime形式
        dates_num = dates.toordinal() + dates.hour/24. + dates.minute/60. + dates.second/3600. -1
        x_all = np.array(list(map(lambda a: (a/50/3600/24+dates_num),np.arange(np.max(Len)))))
        plt.figure()
        kci = 1/60/24 #用于显示得更好看
        for loc in Nright:
            # map映射：lambda函数对Len列表中每个元素进行/50/3600/24+date的计算
            select = [(Len[loc]//4)* i for i in range(4)]
            x = x_all[:int(Len[loc])]
            plt.plot(x, Sig[loc])
            plt.legend(Signame[loc][0])
            plt.xlim(x.min(),x.max()+kci)
            plt.xticks([x[int(i)] for i in select]+[x.max()+kci])
            # x轴刻度的时间格式
            date_format=mdates.DateFormatter('%m-%d %H:%M')

            plt.gca().xaxis.set_major_formatter(date_format)
            # 自动旋转，防止刻度挤在一起重叠
            plt.gcf().autofmt_xdate()
            # 图片保存在data目录的imgs文件夹内，命名规则对应数据文件规则,如: D01.str.png
            fname=savepath+'img/'+ffilename+'/{}.png'.format(filename[loc][0][0])
            fun_save(fname,'img',plt)
            
            #plt.show()
            plt.clf() # 清除图像内存
            plt.gcf().clear() 
            
        # plt.close


        


def signal2(path,ffilename,savepath,Segment,tempplotdata):
    # # 信号显示2
    # for j=[1:148 152:154 159 161:163]
    # 	clf
    # 	plot((1:Len(j))'/50/3600, Sig{j,1})
    # 	legend(Signame(j))
    # 	# xlim([0.5 1])
    # 	title([tstr,'开始记录'])
    # 	pause
    # end



    # 数据调整，将空信号用有用信号代替，之后记得要删除


    ########################################################################
    # 信号分段处理
    # 总通道数164个
    pathstr = os.path.join(path, ffilename)  # 文件路径
    [Sig, Len, Nright, Nerror] = loadSignal20201129(filename, pathstr, segment_length = 720000)
    #Sig list格式 length 164 内容为numpy格式的信号数据 不定长数据 如(720000，)的array
    #Len array格式 length 164 内容为float格式的信号长度 错误或不存在数据长度为0
    #Nright, Nerror array格式 总和length 164 内容为正确/错误数据序号
    
    
    #保存数据
    # fname=savepath+'mat/'+ffilename+'/Sig_Len_Nright.mat'
    # fun_save(fname,'mat',{'Sig':Sig,'Len':Len,'Nright':Nright})
    if Segment<8:
        Index1=list(range(Slen*(Segment-1)+1,Slen*Segment+1))
        # Index2=2*90000*(Segment-1)+1:2:90000*Segment*2;
        DSig=np.zeros((Slen,DSig_shape1))
    else:
        Index1=list(range(Slen*(Segment-1)+1,Slen_+1))
        # Indexl2=2*90000*(Segment-1)+1:2:1439400;
        DSig=np.zeros((Slen,DSig_shape1)); # 为保持数据长度一致，最后一个文件减少3分钟采用数据

    for j in DSigRange: 
        # print(Sig[j].shape)
        DSig[:,j] += Sig[j][Index1[0]-1:Index1[-1]]
            

    # 增加加速度信号换算位移结果
    
    # index_1=[142,143,144,147,145,146] #由于索引从0开始，所有数值已-1
    DDSig=DSig[:,index_1] # 提取加速度A1-A3 和M1

    # JJ=[164,165,166,167,168,169] #由于索引从0开始，所有数值已-1
    """expand DSig"""
    max_jj_col = np.max(JJ)+1
    if max_jj_col>DSig.shape[1]:
        DSig = np.concatenate([DSig,np.zeros((DSig.shape[0],max_jj_col-DSig.shape[1]))],1)
    for j in range(len(JJ)):
        temp=Acc2Disp(DDSig[:,j],SFreq,FF)
        DSig[:,JJ[j]]=temp
        if False:
            # plt.clf()
            
            plt.plot(temp)
            plt.xlim(0,len(temp)-1)
            plt.legend(Signame[JJ[j]])
            fname=savepath+'img/'+ffilename+'/Segment{}.png'.format(Signame[JJ][0][0])
            fun_save(fname,'img',plt)
            #plt.show()
            #pause


    # fname=savepath+'mat/'+ffilename+'/foreDSig'+("%.2d"%Segment)+'.mat'
    # fun_save(fname,'mat',{'DSig':DSig})
    DSig_before = np.round(DSig,12)
    
    # 保存覆写前DDSig，Sigment后数字代表Sigment参数值
    # 对单向应力传感器进行6分钟分段趋势滤除
    for j in range(forcenum):
        DSig[:,j]=Sig_detrend(DSig[:,j],SFreq) # 消除零飘
        if False:
            plt.clf()
            plt.plot(DSig[:,j])
            plt.title(Signame[j])
            plt.show()

            # liu: 由于源代码是手动进行循环，这里进行了原数据的覆盖，后续改代码时需要注意DSig的重声明
    
    # fname=savepath+'mat/'+ffilename+'/afterDSig'+("%.2d"%Segment)+'.mat'
    # fun_save(fname,'mat',{'DSig':DSig})
    [m,n]=np.shape(DSig)
    temp=fftfilter(DSig,SFreq,F2,F4); #	提取波浪和高频合成部分
    Result_Sample=De_Sample(temp,SFreq,F2,F3); # 提取子样
    # 分析锚链张力信号
    # 采用标准差方法计算三一值分析锚链张力信号
    
    Zk=-1
    # index_2=[158,160,161,162] #索引-1
    resultZ=np.zeros((8,1))    # resultZ 8行1列
    for j in index_2:
        Zk=Zk+1
        Z21=fftfilter(DSig[:,j],SFreq,F1,F3) # 提取分析信号，0.01-0.5Hz
        # Result_Sample(j)=De_Sample(temp1,SFreq,F1,F2); # 提取子样
        Z22=fftfilter(Z21,SFreq,F2,F3); # 无慢飘信号
        resultZ[Zk]=2*np.std(Z21,ddof = 1); # 含慢飘
        resultZ[Zk+4]=2*np.std(Z22,ddof = 1); # 无慢飘
        # clf
        # plot(Z21),hold on,plot(Z22)
        # title(num2str(Zk))
        # pause
    # liu: 此段先只考虑完整移植
    # 保留信号子样
    # fname=savepath+'mat/'+ffilename+'/Re'+ttemp[0:10]+("%.2d"%Segment)+'.mat'
    # fun_save(fname,'mat',{'Result_Sample':Result_Sample})
    result=np.zeros((JJ[-1]+1,28))

    # index_3=list(range(112))+list(range(113,148))+[151,152,153,158,160,161,162]+list(range(164,170))#索引-1
    
                
    for j in index_3:
        resulttemp=Sta_irregular(Result_Sample[j])
        # resulttemp=np.array(resulttemp)
        result[j,:]=resulttemp
        #保存用于画图的三种类型的有义值，最大值的三种成分
        for i in range(len(pointname)):
            tempplotdata[j*len(pointname)*8+i*sumSegment+Segment-1]=result[j][int(i/2)*3+1+i%2]
            
            
    result=np.array(result)
    #填入表格
    ws.cell(row=1,column=3).value=ffilename#文件名
    file_hour=int(ffilename[8:-4])
    file_min=int(ffilename[10:-2])
    ws.cell(row=1,column=7).value=str((file_hour+int((30*(Segment-1)+file_min)/60))%24)+':'+str((file_min+((Segment-1)%2)*30)%60)  #starttime
    ws.cell(row=1,column=11).value=str((file_hour+int((30*Segment+file_min)/60))%24)+':'+str((file_min+(Segment%2)*30)%60)  #endtime
    ws.merge_cells('C1:D1')
    ws.merge_cells('G1:H1')
    ws.merge_cells('K1:L1')   #合并单元格
    for i in range(len(Signame)):
        if i in index_3:
            for j in range(result.shape[1]-1):
                ws.cell(row=i*3+4+int(j/9),column=j%9+4).value=result[i][j]
        else:
            for j in range(result.shape[1]-1):
                ws.cell(row=i*3+4+int(j/9),column=j%9+4).value='——'
            # 指定区域单元格居中
    alignment_center = Alignment(horizontal='center', vertical='center')
    ws_area = ws["A1:L522"]
    for i in ws_area:
        for j in i:
            j.alignment = alignment_center
    
    
    excelpath=savepath+'excel/'+ffilename+'/Segment='+str(Segment)+'.xlsx'
    fun_save(excelpath,'xlsx',wb)
    #保存到保存路径中的对应文件夹excel中
    
    
    
    
    # 此处是否需要提供更换字符的接口
    resultD=result[0:10,23]/2
    resultS=result[10:142,23]/2
    resultA=result[indexa,23]/2
    # 保存resultA,resultS,resultD,resultZ
    # fname=savepath+'mat/'+ffilename+'/resultA'+("%.2d"%Segment)+'.mat'
    # fun_save(fname,'mat',{'resultA':resultA})
    # fname=savepath+'mat/'+ffilename+'/resultS'+("%.2d"%Segment)+'.mat'
    # fun_save(fname,'mat',{'resultS':resultS})
    # fname=savepath+'mat/'+ffilename+'/resultD'+("%.2d"%Segment)+'.mat'
    # fun_save(fname,'mat',{'resultD':resultD})
    # fname=savepath+'mat/'+ffilename+'/resultZ'+("%.2d"%Segment)+'.mat'
    # fun_save(fname,'mat',{'resultZ':resultZ})

        

    # val_path = os.path.join('../data','data_val')
    # repro_path = os.path.join('..','matlab_reproduce/')

    # names = os.listdir(val_path)
    # val_data = {}
    # for i in names:
    #     val_data[i.split('.mat')[0]]=scio.loadmat(os.path.join(val_path,i))
    #dict_keys(['DDSig', 'DSig_after', 'DSig_before', 'Len', 'Nright', 'resultZ', 'Result_Sample', 'Sig', 'temp'])

    # 将信号前面补充零值

    #
    # for j=[1:145 149:151 155:157 159]
    # 	len=length(Sig{j:1})
    # 	Sig{j,1}=[zeros(720000-len,1):Sig{j,1}];
    # end
    #

    # 文件合并
    # Sig1=Sig;

    # for j=[1:145 149:151 155:157 159]
    # 	Sig{j:1}=[Sig1{j:1};Sig{j:1}];
    # end
